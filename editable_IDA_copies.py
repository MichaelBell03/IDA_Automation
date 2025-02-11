import os
from openpyxl import Workbook
from xlrd import open_workbook
import pandas as pd
from datetime import datetime

# Directory where the Excel files are located
input_dir = r'C:\Users\c77542\OneDrive - Textron\Attachments\IDA Automation Test Folder'
output_dir = r'C:\Users\c77542\OneDrive - Textron\Attachments\IDA Automation Test Folder'

# Ensure the output directory exists
os.makedirs(output_dir, exist_ok=True)

# Function to load and save a copy of the workbook
def load_and_save_workbook(file_path, output_path):
    # Load .xls file using xlrd
    xls_book = open_workbook(file_path)
    # Create a new .xlsx workbook
    xlsx_book = Workbook()
    sheet = xlsx_book.active
    
    # Read data from columns A, B, and C into lists
    column_a_data = []
    column_b_data = []
    column_c_data = []
    for sheet_index in range(xls_book.nsheets):
        xls_sheet = xls_book.sheet_by_index(sheet_index)
        for row_index in range(xls_sheet.nrows):
            column_a_data.append(xls_sheet.cell_value(row_index, 0))
            column_b_data.append(xls_sheet.cell_value(row_index, 1))
            column_c_data.append(xls_sheet.cell_value(row_index, 2))
    
    for sheet_index in range(xls_book.nsheets):
        xls_sheet = xls_book.sheet_by_index(sheet_index)
        for row_index in range(xls_sheet.nrows):
            for col_index in range(xls_sheet.ncols):
                sheet.cell(row=row_index+1, column=col_index+1, value=xls_sheet.cell_value(row_index, col_index))
    
    # Copy the original data from columns A, B, and C to columns X, Y, and Z
    for row_index in range(len(column_a_data)):
        sheet.cell(row=row_index+1, column=24, value=column_a_data[row_index])
        sheet.cell(row=row_index+1, column=25, value=column_b_data[row_index])
        sheet.cell(row=row_index+1, column=26, value=column_c_data[row_index])
    
    # Remove the default sheet created by openpyxl if it is empty
    if 'Sheet' in xlsx_book.sheetnames and not xlsx_book['Sheet'].max_row:
        xlsx_book.remove(xlsx_book['Sheet'])
    
    # Save the new .xlsx workbook with a distinction in the filename
    base_name, ext = os.path.splitext(output_path)
    xlsx_output_path = f"{base_name}_edit.xlsx"
    xlsx_book.save(xlsx_output_path)

# Iterate through all files in the directory
for file_name in os.listdir(input_dir):
    file_path = os.path.join(input_dir, file_name)
    output_file_path = os.path.join(output_dir, file_name)
    
    # Load and save a copy of the workbook
    load_and_save_workbook(file_path, output_file_path)